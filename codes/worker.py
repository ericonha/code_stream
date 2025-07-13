# models/worker.py
import math
from dataclasses import dataclass, field
import numpy as np
from dataclasses import dataclass
from typing import Optional, List, Tuple

import openpyxl
import pandas as pd
from openpyxl.utils import range_boundaries
from openpyxl import load_workbook
from calendar import monthrange
from typing import List, Tuple, Dict
import io


@dataclass
class Worker:
    id: int
    name: str
    surname: str
    salary: float  # €/hour
    perc_year: List[float]  # % availability per year
    pm_per_ap: float
    years: int  # number of years to cover
    step: float

    hours_available: np.ndarray = field(init=False)
    hours_available_per_month: np.ndarray = field(init=False)
    summed_hours_worked_per_year: np.ndarray = field(init=False)
    summed_hours_worked_total: float = field(init=False)

    assignment_log: Dict[Tuple[int, int], List[Tuple[int, float]]] = field(init=False)
    assignment_log_hours_per_ap: Dict[str, float] = field(init=False)

    def __post_init__(self):
        self.hours_available = np.zeros((self.years, 1))
        self.hours_available_per_month = np.zeros((self.years, 12))
        self.summed_hours_worked_per_year = np.zeros(self.years)
        self.summed_hours_worked_total = 0.0
        self.assignment_log = {}
        self.assignment_log_hours_per_ap = {}

    def set_availability(self, months: List[int], hours_per_year: List[float]):
        for i in range(self.years):
            self.hours_available[i][0] = hours_per_year[i] * (months[i] / 12)
            for m in range(months[i]):
                self.hours_available_per_month[i][m] = 1

    def is_available(self, year: int, month: int, step:float) -> bool:
        return self.hours_available_per_month[year, month] > step

    def print_infos(self):
        print(self.id)
        print(self.name)
        print(self.surname)
        print(self.salary)

def extract_workers_from_dataframe(df: pd.DataFrame, months) -> List[Worker]:
    header_row_clean = df.iloc[0].astype(str).str.strip()
    df_data = df[1:]

    def find_col_index(target: str) -> int:
        for i, col in enumerate(header_row_clean):
            if target.lower() in col.lower():
                return i
        raise ValueError(f"Column containing '{target}' not found.")

    pm_indices = [
        i for i, val in enumerate(header_row_clean)
        if isinstance(val, str) and val.strip() and "max pms per year" in val.lower()
    ]
    perc_year_idx = find_col_index("% per year")
    salary_idx = find_col_index("Salary")
    pm_per_ap_idx = find_col_index("PM/AP")

    workers = []

    for _, row in df_data.iterrows():
        try:
            worker_id = int(row.iloc[0])
            name = str(row.iloc[1])
            surname = str(row.iloc[2])
            step = float(row.iloc[-1]) if not pd.isnull(row.iloc[-1]) else 0.1

            pm_per_year = [float(row.iloc[i]) for i in pm_indices if not pd.isna(row.iloc[i])]
            if not pm_per_year:
                continue  # skip workers with no PMs

            perc_year = [float(row.iloc[perc_year_idx])] * len(pm_per_year)
            salary = float(row.iloc[salary_idx])
            pm_per_ap = float(row.iloc[pm_per_ap_idx]) if not pd.isna(row.iloc[pm_per_ap_idx]) else 10.0

            w = Worker(
                id=worker_id,
                name=name,
                surname=surname,
                salary=salary,
                perc_year=perc_year,
                pm_per_ap=pm_per_ap,
                years=len(pm_per_year),
                step=step
            )
            w.set_availability(
                months=[12] * len(pm_per_year),
                hours_per_year=[pm for pm in pm_per_year]
            )

            index = 0
            for m in months:
                w.hours_available[index] =  w.hours_available[index] * len(months[m])/12
                index += 1

            workers.append(w)

        except Exception as e:
            print(f"Skipping row due to error: {e}")

    return workers

@dataclass
class WorkPackage:
    id: str
    title: str
    required_pm: float
    assigned_worker_name: Optional[str] = None
    assigned_worker_id: Optional[int] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None

    def print_infos(self):
        print(f"[AP {self.id}] '{self.title}' | PM: {self.required_pm} | "
              f"Start: {self.start_date or 'N/A'} | End: {self.end_date or 'N/A'} | "
              f"Worker ID: {self.assigned_worker_id or '—'} | Name: {self.assigned_worker_name or '—'}")

def extract_work_packages_from_dataframe(df: pd.DataFrame, company: str, file_buffer: str) -> List[WorkPackage]:
    aps: List[WorkPackage] = []

    header_row = df.iloc[3].astype(str).str.strip()

    try:
        company_col_idx = next(
            i for i, col in enumerate(header_row)
            if company.lower() in col.lower()
        )
    except StopIteration:
        print(f"❌ Company '{company}' not found in the Excel header row.")
        return []

    # ✅ Use the buffer to extract month start/end
    month_col_map = extract_ap_date_ranges(file_buffer, company)
    index_ap = 0

    for _, row in df.iloc[4:].iterrows():
        ap_id_raw = row.iloc[1]

        if pd.isna(ap_id_raw):
            continue

        ap_id_str = str(ap_id_raw).strip()

        if len(month_col_map) == index_ap:
            break

        try:
            if float(ap_id_str).is_integer():
                continue
        except ValueError:
            pass

        title = str(row.iloc[2]).strip() if not pd.isna(row.iloc[2]) else ""

        try:
            required_pm = float(row.iloc[company_col_idx])
        except Exception:
            required_pm = 0.0

        if required_pm <= 0 or math.isnan(required_pm):
            continue

        assigned_worker = row.iloc[6] if not pd.isna(row.iloc[6]) else None

        ap = WorkPackage(
            id=ap_id_str,
            title=title,
            required_pm=required_pm,
            assigned_worker_name=str(assigned_worker).strip() if assigned_worker else None,
            start_date=month_col_map[index_ap][0],
            end_date=month_col_map[index_ap][1],
        )
        index_ap += 1
        aps.append(ap)

    return aps

MONTH_MAP = {
    "jan": 1, "januar": 1,
    "feb": 2, "februar": 2,
    "mar": 3, "märz": 3, "maerz": 3, "mrz": 3,
    "apr": 4, "april": 4,
    "mai": 5,
    "jun": 6, "juni": 6,
    "jul": 7, "juli": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "okt": 10, "oct": 10, "oktober": 10,
    "nov": 11, "november": 11,
    "dez": 12, "dec": 12, "dezember": 12
}

def month_str_to_number(month_str: str) -> int:
    key = str(month_str).strip().lower()
    if key in MONTH_MAP:
        return MONTH_MAP[key]
    else:
        raise ValueError(f"Unknown month format: '{month_str}'")

def get_merged_cell_value(ws, row: int, col: int):
    """
    Retorna o valor da célula principal se a (row, col) estiver dentro de uma célula mesclada.
    Caso contrário, retorna None.
    """
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return ws.cell(row=min_row, column=min_col).value  # valor da célula principal
    return None  # Não está em uma célula mesclada

def extract_ap_date_ranges(filename: str, company: str) -> List[Tuple[str, str]]:
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active

    max_col = ws.max_column
    max_row = ws.max_row
    header_row_idx = 4
    month_row_idx = 4
    year_row_idx = 3
    data_start_row = 6

    # Find 'Summe' column
    date_start_col = None
    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row_idx, column=col).value
        if val and isinstance(val, str) and "summe" in val.lower():
            date_start_col = col + 1
            break
    if date_start_col is None:
        print("❌ Could not find 'Summe' column.")
        return []

    # Find company column
    company_col = None
    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row_idx, column=col).value
        if val and isinstance(val, str) and company.lower() in val.lower():
            company_col = col
            break
    if company_col is None:
        print(f"❌ Company '{company}' not found.")
        return []

    # Determine comparator color
    reference_color = ws.cell(row=data_start_row, column=date_start_col).fill.start_color.index
    comparator_color = 8 if reference_color > 2 else 1

    date_ranges: List[Tuple[str, str]] = []

    for row in range(data_start_row, max_row + 1):
        ap_id_cell = ws.cell(row=row, column=2).value
        if not ap_id_cell or str(ap_id_cell).strip() == "":
            continue  # skip empty rows

        # Skip whole-number IDs like 1, 2, ...
        try:
            if float(ap_id_cell).is_integer():
                continue
        except:
            pass

        # Skip if company effort is 0 or invalid
        try:
            effort = float(ws.cell(row=row, column=company_col).value or 0.0)
            if effort <= 0:
                continue
        except:
            continue

        # Find black-filled columns (continuous block)
        black_cols = []
        for col in range(date_start_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill.start_color.index == comparator_color:
                black_cols.append(col)
            elif black_cols:
                break  # stop after black block ends

        if not black_cols:
            continue

        first_col = black_cols[0]
        last_col = black_cols[-1]

        if row == 25:
            print("ok")

        try:
            month1_str = str(ws.cell(row=month_row_idx, column=first_col).value)
            month1 = month_str_to_number(month1_str)
            year_raw = get_merged_cell_value(ws, year_row_idx, first_col)
            if year_raw is None:
                raise ValueError(f"Couldn't find a year value for column {first_col}")
            year1 = int(year_raw)
            start_date = f"01.{month1:02d}.{year1}"
        except:
            start_date = None

        try:
            month2_str = str(ws.cell(row=month_row_idx, column=last_col).value)
            month2 = month_str_to_number(month2_str)
            year_raw = get_merged_cell_value(ws, year_row_idx, last_col)
            if year_raw is None:
                raise ValueError(f"Couldn't find a year value for column {last_col}")
            year2 = int(year_raw)
            last_day = int(monthrange(2026, month2)[1])
            end_date = f"{last_day:02d}.{month2:02d}.{year2:02d}"
        except Exception as e:
            end_date = None
            print(e)

        if start_date and end_date:
            date_ranges.append((start_date, end_date))

    return date_ranges

def month_per_year(df: pd.DataFrame):
    header_row_year = df.iloc[2].copy()
    header_row_month = df.iloc[3]

    # Encontrar índice da coluna "Summe"
    summe_col_idx = None
    for i, val in enumerate(header_row_month):
        if isinstance(val, str) and "summe" in val.lower():
            summe_col_idx = i
            break

    if summe_col_idx is None:
        raise ValueError("❌ Coluna 'Summe' não encontrada.")

    # Preencher os NaNs com o último ano válido para lidar com células mescladas
    header_row_year.iloc[summe_col_idx + 1:] = header_row_year.iloc[summe_col_idx + 1:].ffill()

    months_by_year = {}
    for col in range(summe_col_idx + 1, len(df.columns)):
        year = header_row_year[col]
        month = header_row_month[col]

        if pd.isna(year) or pd.isna(month):
            continue

        year = str(year).strip()
        month = str(month).strip()

        if year not in months_by_year:
            months_by_year[year] = []

        months_by_year[year].append(month)

    return months_by_year

def is_in_project(uploaded_ap_file, company):
    if uploaded_ap_file is None:
        return 1

    df = pd.read_excel(uploaded_ap_file, header=None)

    header_row = df.iloc[3].astype(str).str.strip()

    try:
        company_col_idx = next(
            i for i, col in enumerate(header_row)
            if company.lower() in col.lower()
        )
    except StopIteration:
        print(f"❌ Company '{company}' not found in the Excel header row.")
        return []

    return None
